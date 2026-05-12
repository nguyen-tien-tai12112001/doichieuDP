"""
Export module for generating Excel reports.
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import io


def format_currency(value):
    """Format value as currency."""
    if pd.isna(value):
        return 0
    return f"{value:,.0f}"


def format_percentage(value):
    """Format value as percentage."""
    if pd.isna(value):
        return "0%"
    return f"{value:.2%}"


def apply_header_style(worksheet, header_row: int):
    """Apply styling to header row."""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in worksheet[header_row]:
        if cell.value:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def apply_data_style(worksheet, start_row: int, end_row: int):
    """Apply styling to data rows."""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in worksheet.iter_rows(min_row=start_row, max_row=end_row):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")


def export_to_excel(
    comparison_df: pd.DataFrame,
    summary_branch: pd.DataFrame,
    summary_cust_type: pd.DataFrame,
    summary_product: pd.DataFrame,
    outliers_df: pd.DataFrame,
    recommendations_df: pd.DataFrame = None,
    driver_analysis_df: pd.DataFrame = None,
    segment_summary_df: pd.DataFrame = None,
    alert_summary_df: pd.DataFrame = None,
    prediction_df: pd.DataFrame = None,
    market_context: list[str] = None,
    capital_strategy: list[dict] = None,
    ai_insights: list[str] = None,
    filename: str = None
) -> bytes:
    """
    Export all data to Excel with multiple sheets.
    Optimized for speed with minimal styling.

    Args:
        comparison_df: Detailed comparison data
        summary_branch: Summary by branch
        summary_cust_type: Summary by customer type
        summary_product: Summary by product group
        outliers_df: Outlier data
        recommendations_df: Action recommendation data (optional)
        driver_analysis_df: Driver analysis data (optional)
        segment_summary_df: Segment summary data (optional)
        alert_summary_df: Alert summary data (optional)
        prediction_df: Forecast data (optional)
        market_context: Market context narrative (optional)
        capital_strategy: List of strategy guidance (optional)
        ai_insights: AI insight list (optional)
        filename: Output filename (optional)

    Returns:
        Excel file as bytes
    """
    output = io.BytesIO() if not filename else None
    writer_path = output if output else filename

    # Shared header style, always available for all sheets.
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    def _write_text_sheet(sheet_name: str, rows: list[str]):
        if not rows:
            return
        df_export = pd.DataFrame({'Nội dung': rows})
        df_export.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        for col_num in range(1, len(df_export.columns) + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 100

    with pd.ExcelWriter(writer_path, engine='openpyxl') as writer:
        # Sheet 1: Detailed customers
        if not comparison_df.empty:
            df_export = comparison_df.copy()
            # Format numeric columns efficiently
            for col in ['TOTAL_T1', 'TOTAL_T2', 'DELTA']:
                if col in df_export.columns:
                    df_export[col] = df_export[col].apply(lambda x: f"{x:,.0f}" if pd.notna(x) else "0")
            
            df_export.to_excel(writer, sheet_name='Chi Tiết Khách Hàng', index=False)
            
            # Minimal header styling
            ws = writer.sheets['Chi Tiết Khách Hàng']
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            
            # Set fixed column widths for speed
            for col_num, column_title in enumerate(df_export.columns, 1):
                col_letter = get_column_letter(col_num)
                ws.column_dimensions[col_letter].width = 15
        
        # Sheet 2: Executive Summary
        executive_summary = []
        if not comparison_df.empty:
            total_customers = len(comparison_df)
            total_t1 = float(comparison_df['TOTAL_T1'].sum())
            total_t2 = float(comparison_df['TOTAL_T2'].sum())
            total_delta = float(comparison_df['DELTA'].sum())
            executive_summary.extend([
                f'Tổng khách hàng so sánh: {total_customers}',
                f'Tổng T1: {total_t1:,.0f}',
                f'Tổng T2: {total_t2:,.0f}',
                f'Tổng DELTA: {total_delta:,.0f}',
                f'Tỉ lệ tăng trưởng tổng: {((total_delta / max(abs(total_t1), 1)) * 100):.2f}%'
            ])
        if market_context:
            executive_summary.append('---')
            executive_summary.append('Bối cảnh thị trường:')
            executive_summary.extend(market_context)
        if capital_strategy:
            executive_summary.append('---')
            executive_summary.append('Chiến lược huy động vốn:')
            for strategy in capital_strategy:
                executive_summary.append(f"{strategy['title']}: {strategy['description']}")
        if ai_insights:
            executive_summary.append('---')
            executive_summary.append('AI insights tổng hợp:')
            executive_summary.extend(ai_insights)

        if executive_summary:
            _write_text_sheet('Executive Summary', executive_summary)

        # Sheet 3: Summary by branch
        if not summary_branch.empty:
            df_export = summary_branch.copy()
            for col in ['TONG_T1', 'TONG_T2', 'TONG_DELTA']:
                if col in df_export.columns:
                    df_export[col] = df_export[col].apply(lambda x: f"{x:,.0f}" if pd.notna(x) else "0")
            if 'TY_LE_TANG_TRUONG' in df_export.columns:
                df_export['TY_LE_TANG_TRUONG'] = df_export['TY_LE_TANG_TRUONG'].apply(lambda x: f"{x:.2%}" if pd.notna(x) else "0%")
            
            df_export.to_excel(writer, sheet_name='Theo Chi Nhánh', index=False)
            
            ws = writer.sheets['Theo Chi Nhánh']
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            for col_num in range(1, len(df_export.columns) + 1):
                ws.column_dimensions[get_column_letter(col_num)].width = 18
        
        # Sheet 3: Summary by customer type
        if not summary_cust_type.empty:
            df_export = summary_cust_type.copy()
            for col in ['TONG_T1', 'TONG_T2', 'TONG_DELTA']:
                if col in df_export.columns:
                    df_export[col] = df_export[col].apply(lambda x: f"{x:,.0f}" if pd.notna(x) else "0")
            
            df_export.to_excel(writer, sheet_name='Theo Phân Khúc', index=False)
            
            ws = writer.sheets['Theo Phân Khúc']
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            for col_num in range(1, len(df_export.columns) + 1):
                ws.column_dimensions[get_column_letter(col_num)].width = 18
        
        # Sheet 4: Summary by product group
        if not summary_product.empty:
            df_export = summary_product.copy()
            for col in ['TONG_T1', 'TONG_T2', 'TONG_DELTA']:
                if col in df_export.columns:
                    df_export[col] = df_export[col].apply(lambda x: f"{x:,.0f}" if pd.notna(x) else "0")
            
            df_export.to_excel(writer, sheet_name='Theo Sản Phẩm', index=False)
            
            ws = writer.sheets['Theo Sản Phẩm']
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            for col_num in range(1, len(df_export.columns) + 1):
                ws.column_dimensions[get_column_letter(col_num)].width = 18
        
        # Sheet 5: Outliers
        if not outliers_df.empty:
            df_export = outliers_df.copy()
            for col in ['TOTAL_T1', 'TOTAL_T2', 'DELTA']:
                if col in df_export.columns:
                    df_export[col] = df_export[col].apply(lambda x: f"{x:,.0f}" if pd.notna(x) else "0")
            
            outlier_cols = ['MA_CN', 'MA_KH', 'TEN_KH', 'CUST_TYPE_NAME', 'TOTAL_T1', 'TOTAL_T2', 'DELTA', 'BIEN_DONG']
            df_export = df_export[[col for col in outlier_cols if col in df_export.columns]]
            
            df_export.to_excel(writer, sheet_name='Bất Thường', index=False)
            
            ws = writer.sheets['Bất Thường']
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            for col_num in range(1, len(df_export.columns) + 1):
                ws.column_dimensions[get_column_letter(col_num)].width = 15

        # Sheet 6: Action recommendations
        if recommendations_df is not None and not recommendations_df.empty:
            df_export = recommendations_df.copy()
            for col in ['TOTAL_T1', 'TOTAL_T2', 'DELTA', 'ABS_DELTA']:
                if col in df_export.columns:
                    df_export[col] = df_export[col].apply(lambda x: f"{x:,.0f}" if pd.notna(x) else "0")

            cols = [
                'MUC_UU_TIEN', 'NHOM_KHUYEN_NGHI', 'HANH_DONG_GOI_Y', 'LY_DO',
                'MA_CN', 'MA_KH', 'TEN_KH', 'CUST_TYPE_NAME',
                'TOTAL_T1', 'TOTAL_T2', 'DELTA', 'BIEN_DONG', 'ABS_DELTA', 'DIEM_RUI_RO'
            ]
            df_export = df_export[[col for col in cols if col in df_export.columns]]
            df_export.to_excel(writer, sheet_name='Khuyến Nghị', index=False)

            ws = writer.sheets['Khuyến Nghị']
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            for col_num in range(1, len(df_export.columns) + 1):
                ws.column_dimensions[get_column_letter(col_num)].width = 18

        # Sheet 7: Market context
        if market_context:
            _write_text_sheet('Market Context', market_context)

        # Sheet 8: Capital strategy
        if capital_strategy:
            strategy_lines = []
            for item in capital_strategy:
                strategy_lines.append(f"{item['title']} - {item['description']}")
            _write_text_sheet('Capital Strategy', strategy_lines)

        # Sheet 9: AI Insights
        if ai_insights:
            _write_text_sheet('AI Insights', ai_insights)

        # Sheet 10: Forecast
        if prediction_df is not None and not prediction_df.empty:
            df_export = prediction_df.copy()
            df_export.to_excel(writer, sheet_name='Dự Báo', index=False)
            ws = writer.sheets['Dự Báo']
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            for col_num in range(1, len(df_export.columns) + 1):
                ws.column_dimensions[get_column_letter(col_num)].width = 18

        # Sheet 7: Driver analysis
        if driver_analysis_df is not None and not driver_analysis_df.empty:
            df_export = driver_analysis_df.copy()
            if 'TONG_DELTA' in df_export.columns:
                df_export['TONG_DELTA'] = df_export['TONG_DELTA'].apply(lambda x: f"{x:,.0f}" if pd.notna(x) else "0")
            for col in ['DONG_GOP_RONG_%', 'DONG_GOP_ABS_%']:
                if col in df_export.columns:
                    df_export[col] = df_export[col].apply(lambda x: f"{x:.2f}%" if pd.notna(x) else "0.00%")

            cols = ['NHOM_PHAN_TICH', 'NHOM', 'TONG_DELTA', 'DONG_GOP_RONG_%', 'DONG_GOP_ABS_%', 'XU_HUONG']
            df_export = df_export[[col for col in cols if col in df_export.columns]]
            df_export.to_excel(writer, sheet_name='Động Lực Biến Động', index=False)

            ws = writer.sheets['Động Lực Biến Động']
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            for col_num in range(1, len(df_export.columns) + 1):
                ws.column_dimensions[get_column_letter(col_num)].width = 20

        # Sheet 8: Segment summary
        if segment_summary_df is not None and not segment_summary_df.empty:
            df_export = segment_summary_df.copy()
            for col in ['TONG_T1', 'TONG_T2', 'TONG_DELTA', 'TY_LE_TANG_TRUONG', 'DIEM_RUI_RO_TRUNG_BINH']:
                if col in df_export.columns:
                    if col == 'TY_LE_TANG_TRUONG':
                        df_export[col] = df_export[col].apply(lambda x: f"{x:.2f}%" if pd.notna(x) else "0.00%")
                    else:
                        df_export[col] = df_export[col].apply(lambda x: f"{x:,.0f}" if pd.notna(x) else "0")

            df_export.to_excel(writer, sheet_name='Phân Khúc Nâng Cao', index=False)
            ws = writer.sheets['Phân Khúc Nâng Cao']
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            for col_num in range(1, len(df_export.columns) + 1):
                ws.column_dimensions[get_column_letter(col_num)].width = 20

        # Sheet 9: Alert summary
        if alert_summary_df is not None and not alert_summary_df.empty:
            df_export = alert_summary_df.copy()
            df_export.to_excel(writer, sheet_name='Cảnh Báo', index=False)
            ws = writer.sheets['Cảnh Báo']
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            for col_num in range(1, len(df_export.columns) + 1):
                ws.column_dimensions[get_column_letter(col_num)].width = 22
    
    if output:
        output.seek(0)
        return output.getvalue()
    
    return None
