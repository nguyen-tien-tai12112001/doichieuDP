"""
Main Streamlit application for deposit comparison system.
"""
import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import hashlib
from pathlib import Path
import os
import io
import json
from datetime import datetime
import tempfile
from typing import Dict, List, Tuple
import pickle
import sqlite3
from contextlib import contextmanager

# Import custom modules
from validators import validate_all_files, ValidationError
from loader import load_and_normalize_csv, filter_valid_data, load_branch_data
from aggregator import group_by_customer, aggregate_pair
from compare_engine import merge_and_compare, process_all_branches
from summary_engine import (
    summary_by_branch, summary_by_customer_type, summary_by_product_group,
    summary_by_change_type, get_dp_type_mapping
)
from outlier_engine import detect_outliers_iqr, get_outliers_summary, analyze_outliers
from exporter import export_to_excel


CANONICAL_COLUMNS = ['MA_KH', 'TEN_KH', 'DP_TYPE_CODE', 'CURRENT_BALANCE', 'CUST_TYPE_NAME']
DATABASE_FILE = Path('app_database.db')

SEGMENT_BUCKETS = [
    ('<50M', 0, 50_000_000),
    ('50-200M', 50_000_000, 200_000_000),
    ('200-500M', 200_000_000, 500_000_000),
    ('>500M', 500_000_000, float('inf')),
]
ALERT_BRANCH_DROP_RATE = -0.05
ALERT_BRANCH_DELTA_THRESHOLD = 100_000_000
ALERT_PRODUCT_DELTA_THRESHOLD = 150_000_000
ALERT_OUTLIER_THRESHOLD = 5


@contextmanager
def get_db_connection():
    """Context manager for database connections."""
    conn = sqlite3.connect(DATABASE_FILE)
    conn.row_factory = sqlite3.Row  # Enable column access by name
    try:
        yield conn
    finally:
        conn.close()


def init_database():
    """Initialize database tables if they don't exist."""
    with get_db_connection() as conn:
        cursor = conn.cursor()

        # History table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS comparison_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                timestamp TEXT NOT NULL,
                cache_key TEXT UNIQUE,
                t1_input_files TEXT,
                t2_input_files TEXT,
                t1_files INTEGER,
                t2_files INTEGER,
                t1_date TEXT,
                t2_date TEXT,
                branches INTEGER,
                customers INTEGER,
                total_t1 REAL,
                total_t2 REAL,
                total_delta REAL,
                high_risk_customers INTEGER,
                cache_hit TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # Processed data cache table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS processed_data_cache (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                cache_key TEXT UNIQUE NOT NULL,
                data_json TEXT NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                last_accessed TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # User settings table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS user_settings (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                setting_key TEXT UNIQUE NOT NULL,
                setting_value TEXT,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        conn.commit()


def save_history_to_db(entry: Dict[str, str]) -> None:
    """Save comparison history to database."""
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('''
            INSERT OR REPLACE INTO comparison_history
            (timestamp, cache_key, t1_input_files, t2_input_files, t1_files, t2_files,
             t1_date, t2_date, branches, customers, total_t1, total_t2, total_delta,
             high_risk_customers, cache_hit)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            entry.get('timestamp'),
            entry.get('cache_key'),
            entry.get('t1_input_files'),
            entry.get('t2_input_files'),
            entry.get('t1_files'),
            entry.get('t2_files'),
            entry.get('t1_date'),
            entry.get('t2_date'),
            entry.get('branches'),
            entry.get('customers'),
            entry.get('total_t1'),
            entry.get('total_t2'),
            entry.get('total_delta'),
            entry.get('high_risk_customers'),
            entry.get('cache_hit')
        ))
        conn.commit()


def load_history_from_db(limit: int = 20) -> pd.DataFrame:
    """Load recent processing history from database."""
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('''
            SELECT timestamp, cache_key, t1_input_files, t2_input_files, t1_files, t2_files,
                   t1_date, t2_date, branches, customers, total_t1, total_t2, total_delta,
                   high_risk_customers, cache_hit
            FROM comparison_history
            ORDER BY timestamp DESC
            LIMIT ?
        ''', (limit,))
        rows = cursor.fetchall()

        if not rows:
            return pd.DataFrame()

        # Convert to DataFrame
        data = []
        for row in rows:
            data.append({
                'timestamp': row['timestamp'],
                'cache_key': row['cache_key'],
                't1_input_files': row['t1_input_files'],
                't2_input_files': row['t2_input_files'],
                't1_files': row['t1_files'],
                't2_files': row['t2_files'],
                't1_date': row['t1_date'],
                't2_date': row['t2_date'],
                'branches': row['branches'],
                'customers': row['customers'],
                'total_t1': row['total_t1'],
                'total_t2': row['total_t2'],
                'total_delta': row['total_delta'],
                'high_risk_customers': row['high_risk_customers'],
                'cache_hit': row['cache_hit']
            })

        return pd.DataFrame(data)


def save_processed_data_to_db(cache_key: str, processed_data: Dict) -> None:
    """Save processed data to database cache."""
    try:
        # Serialize the data using pickle (better for pandas DataFrames)
        import pickle
        data_blob = pickle.dumps(processed_data)

        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                INSERT OR REPLACE INTO processed_data_cache
                (cache_key, data_json, last_accessed)
                VALUES (?, ?, CURRENT_TIMESTAMP)
            ''', (cache_key, data_blob))
            conn.commit()
    except Exception as e:
        st.warning(f"Không thể lưu cache vào database: {e}")


def load_processed_data_from_db(cache_key: str) -> Dict:
    """Load processed data from database cache."""
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                SELECT data_json FROM processed_data_cache
                WHERE cache_key = ?
            ''', (cache_key,))
            row = cursor.fetchone()

            if row:
                # Update last accessed time
                cursor.execute('''
                    UPDATE processed_data_cache
                    SET last_accessed = CURRENT_TIMESTAMP
                    WHERE cache_key = ?
                ''', (cache_key,))
                conn.commit()

                # Deserialize the data - handle both old JSON and new pickle formats
                data = row['data_json']
                import pickle

                # Try pickle first (new format)
                try:
                    return pickle.loads(data)
                except (TypeError, pickle.UnpicklingError):
                    # Fall back to JSON (old format) - but this will return strings, not DataFrames
                    # So we'll return empty dict to force re-processing
                    st.info("Cache cũ không tương thích, sẽ tạo cache mới.")
                    return {}
    except Exception as e:
        st.warning(f"Không thể tải cache từ database: {e}")

    return {}


def cleanup_old_cache(days: int = 30) -> None:
    """Clean up old cache entries older than specified days."""
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                DELETE FROM processed_data_cache
                WHERE last_accessed < datetime('now', '-{} days')
            '''.format(days))
            deleted_count = cursor.rowcount
            conn.commit()

            if deleted_count > 0:
                st.info(f"Đã dọn dẹp {deleted_count} cache entries cũ")
    except Exception as e:
        st.warning(f"Lỗi khi dọn dẹp cache: {e}")


# Initialize database on startup
init_database()

SEGMENT_BUCKETS = [
    ('<50M', 0, 50_000_000),
    ('50-200M', 50_000_000, 200_000_000),
    ('200-500M', 200_000_000, 500_000_000),
    ('>500M', 500_000_000, float('inf')),
]
ALERT_BRANCH_DROP_RATE = -0.05
ALERT_BRANCH_DELTA_THRESHOLD = 100_000_000
ALERT_PRODUCT_DELTA_THRESHOLD = 150_000_000
ALERT_OUTLIER_THRESHOLD = 5


# Configure page
st.set_page_config(
    page_title="Hệ Thống So Sánh Dữ Liệu Tiền Gửi",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Add enhanced CSS styling
st.markdown("""
<style>
    .main-header {
        color: #1f77b4;
        font-size: 2.8rem;
        font-weight: bold;
        margin-bottom: 10px;
        text-align: center;
    }
    .sub-header {
        color: #666;
        font-size: 1.1rem;
        margin-bottom: 30px;
        text-align: center;
    }
    .step-guide {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        padding: 20px;
        border-radius: 10px;
        margin-bottom: 20px;
    }
    .step-guide h3 {
        margin-top: 0;
    }
    .info-box {
        background-color: #e7f3ff;
        border-left: 4px solid #2196F3;
        padding: 15px;
        border-radius: 5px;
        margin: 15px 0;
    }
    .success-box {
        background-color: #e8f5e9;
        border-left: 4px solid #4CAF50;
        padding: 15px;
        border-radius: 5px;
        margin: 15px 0;
    }
    .warning-box {
        background-color: #fff3e0;
        border-left: 4px solid #FF9800;
        padding: 15px;
        border-radius: 5px;
        margin: 15px 0;
    }
    .metric-card {
        background: linear-gradient(135deg, #f5f7fa 0%, #c3cfe2 100%);
        padding: 20px;
        border-radius: 10px;
        margin: 10px 0;
        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
    }
    .stTabs [data-baseweb="tab-list"] button {
        font-size: 16px;
        font-weight: 600;
    }
</style>
""", unsafe_allow_html=True)

def init_session_state() -> None:
    """Initialize session state keys used by the app."""
    defaults = {
        'processed_data': None,
        'validation_result': None,
        'upload_context': None,
        'processing_cache': {},
    }

    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value


def get_current_column_mapping() -> Dict[str, str]:
    """Read current mapping values from widget state."""
    mapping = {}
    for canonical in CANONICAL_COLUMNS:
        key = f'map_{canonical}'
        if key not in st.session_state:
            st.session_state[key] = canonical
        mapping[canonical] = st.session_state[key].strip() or canonical
    return mapping


def generate_template_csv(column_mapping: Dict[str, str]) -> bytes:
    """Generate downloadable CSV template using current column mapping."""
    sample_df = pd.DataFrame([
        {
            'MA_KH': 'CUST001',
            'TEN_KH': 'NGUYEN VAN A',
            'DP_TYPE_CODE': '010',
            'CURRENT_BALANCE': 150000000,
            'CUST_TYPE_NAME': 'CA_NHAN',
        },
        {
            'MA_KH': 'CUST002',
            'TEN_KH': 'CONG TY ABC',
            'DP_TYPE_CODE': '020',
            'CURRENT_BALANCE': 520000000,
            'CUST_TYPE_NAME': 'PHAP_NHAN',
        },
    ])

    rename_map = {canonical: column_mapping.get(canonical, canonical) for canonical in CANONICAL_COLUMNS}
    sample_df = sample_df.rename(columns=rename_map)
    return sample_df.to_csv(index=False).encode('utf-8-sig')


def save_uploaded_files(files, temp_dir: str) -> List[str]:
    """Persist uploaded files to temporary directory."""
    paths = []
    for file in files:
        path = os.path.join(temp_dir, file.name)
        with open(path, 'wb') as f:
            f.write(file.getbuffer())
        paths.append(path)
    return paths


def collect_file_metadata(files) -> List[Dict[str, str]]:
    """Collect filename and content hash for cache/history."""
    metadata = []
    for file in files:
        raw = file.getvalue()
        digest = hashlib.sha256(raw).hexdigest()
        metadata.append({'name': file.name, 'sha256': digest})
    return sorted(metadata, key=lambda x: x['name'])


def build_cache_key(
    t1_meta: List[Dict[str, str]],
    t2_meta: List[Dict[str, str]],
    column_mapping: Dict[str, str],
) -> str:
    """Build a deterministic cache key from inputs and mapping."""
    payload = {
        't1': t1_meta,
        't2': t2_meta,
        'column_mapping': column_mapping,
    }
    raw = json.dumps(payload, sort_keys=True, ensure_ascii=True)
    return hashlib.sha256(raw.encode('utf-8')).hexdigest()


def append_history(entry: Dict[str, str]) -> None:
    """Append one processing record to database."""
    save_history_to_db(entry)


def save_processed_data(cache_key: str, processed_data: Dict) -> None:
    """Save processed data to database cache."""
    save_processed_data_to_db(cache_key, processed_data)


def load_processed_data(cache_key: str) -> Dict:
    """Load processed data from database cache."""
    return load_processed_data_from_db(cache_key)


def load_history(limit: int = 20) -> pd.DataFrame:
    """Load recent processing history from database."""
    return load_history_from_db(limit)


def render_friendly_error(error_text: str) -> None:
    """Display user-friendly validation errors with suggestions."""
    st.error(f"❌ {error_text}")
    st.markdown(
        """
        <div class="warning-box">
            <b>Gợi ý xử lý nhanh:</b><br>
            1) Kiểm tra tên file có dạng <code>{MA_CN}_dp01_yyyymmdd.csv</code><br>
            2) Tải template CSV mẫu để đối chiếu cấu trúc cột<br>
            3) Nếu cột nguồn khác tên chuẩn, chỉnh lại ở mục <b>Mapping Cột</b><br>
            4) Re-run bước 2: Kiểm tra dữ liệu
        </div>
        """,
        unsafe_allow_html=True,
    )


def _safe_quantile(series: pd.Series, q: float, default_value: float) -> float:
    """Get quantile safely with fallback for empty/invalid series."""
    if series is None or series.empty:
        return default_value

    value = series.quantile(q)
    if pd.isna(value):
        return default_value

    return float(value)


def build_action_recommendations(comparison_df: pd.DataFrame) -> tuple[pd.DataFrame, dict]:
    """Build rule-based customer action recommendations."""
    if comparison_df.empty:
        return pd.DataFrame(), {
            'risk_threshold': 0,
            'growth_threshold': 0,
            'new_open_threshold': 0,
        }

    negative_abs = comparison_df.loc[comparison_df['DELTA'] < 0, 'DELTA'].abs()
    positive_vals = comparison_df.loc[comparison_df['DELTA'] > 0, 'DELTA']
    new_open_vals = comparison_df.loc[comparison_df['BIEN_DONG'] == 'MO_MOI', 'TOTAL_T2']

    risk_threshold = max(50_000_000, _safe_quantile(negative_abs, 0.75, 50_000_000))
    growth_threshold = max(50_000_000, _safe_quantile(positive_vals, 0.75, 50_000_000))
    new_open_threshold = max(100_000_000, _safe_quantile(new_open_vals, 0.75, 100_000_000))

    rows = []
    for row in comparison_df.itertuples(index=False):
        priority = None
        action = None
        recommendation_group = None
        reason = None

        if row.BIEN_DONG == 'TAT_TOAN' and row.TOTAL_T1 > 0:
            priority = 'RAT_CAO'
            recommendation_group = 'GIU_CHAN'
            action = 'Lien he trong 48h, tim nguyen nhan tat toan va de xuat tai tuc.'
            reason = f'Tat toan toan bo {row.TOTAL_T1:,.0f}.'
        elif row.DELTA <= -1.5 * risk_threshold:
            priority = 'RAT_CAO'
            recommendation_group = 'GIU_CHAN'
            action = 'Uu tien cham soc ngay, de xuat goi lai suat/uu dai chuyen biet.'
            reason = f'Giam rat manh {row.DELTA:,.0f}, vuot nguong canh bao cao.'
        elif row.DELTA <= -risk_threshold:
            priority = 'CAO'
            recommendation_group = 'GIU_CHAN'
            action = 'Goi cham soc chu dong, xac minh nhu cau va de xuat san pham phu hop.'
            reason = f'Giam manh {row.DELTA:,.0f}, vuot nguong canh bao.'
        elif row.BIEN_DONG == 'MO_MOI' and row.TOTAL_T2 >= new_open_threshold:
            priority = 'TRUNG_BINH'
            recommendation_group = 'ONBOARDING'
            action = 'Onboarding va cross-sell san pham phu tro trong 7 ngay dau.'
            reason = f'Mo moi gia tri lon {row.TOTAL_T2:,.0f}.'
        elif row.TOTAL_T2 >= 200_000_000 and row.DELTA > 0:
            priority = 'TRUNG_BINH'
            recommendation_group = 'MO_RONG'
            action = 'Xem xet upsell/cross-sell san pham phu hop cho khach hang gia tri cao.'
            reason = f'Khach tang va co du lon {row.TOTAL_T2:,.0f}.'
        elif row.DELTA <= -risk_threshold and row.TOTAL_T2 >= 100_000_000:
            priority = 'CAO'
            recommendation_group = 'GIU_CHAN'
            action = 'Kiem tra chinh sach phi va uu dai giu chan khach hang gia tri cao.'
            reason = f'Giam manh {row.DELTA:,.0f} nhung van co du lon {row.TOTAL_T2:,.0f}.'
        elif row.DELTA >= growth_threshold:
            priority = 'TRUNG_BINH'
            recommendation_group = 'MO_RONG'
            action = 'Duy tri da tang truong, de xuat goi gia tri cao hon/VIP retention.'
            reason = f'Tang manh {row.DELTA:,.0f}.'

        if priority is not None:
            rows.append({
                'MUC_UU_TIEN': priority,
                'NHOM_KHUYEN_NGHI': recommendation_group,
                'HANH_DONG_GOI_Y': action,
                'LY_DO': reason,
                'MA_CN': row.MA_CN,
                'MA_KH': row.MA_KH,
                'TEN_KH': row.TEN_KH,
                'CUST_TYPE_NAME': row.CUST_TYPE_NAME,
                'TOTAL_T1': row.TOTAL_T1,
                'TOTAL_T2': row.TOTAL_T2,
                'DELTA': row.DELTA,
                'BIEN_DONG': row.BIEN_DONG,
                'ABS_DELTA': abs(row.DELTA),
                'DIEM_RUI_RO': getattr(row, 'CHURN_SCORE', 0),
            })

    if not rows:
        return pd.DataFrame(), {
            'risk_threshold': risk_threshold,
            'growth_threshold': growth_threshold,
            'new_open_threshold': new_open_threshold,
        }

    recommendations_df = pd.DataFrame(rows)
    priority_order = {'RAT_CAO': 0, 'CAO': 1, 'TRUNG_BINH': 2}
    recommendations_df['priority_rank'] = recommendations_df['MUC_UU_TIEN'].map(priority_order).fillna(9)
    recommendations_df = recommendations_df.sort_values(
        ['priority_rank', 'ABS_DELTA'],
        ascending=[True, False],
    ).drop(columns=['priority_rank']).reset_index(drop=True)

    return recommendations_df, {
        'risk_threshold': risk_threshold,
        'growth_threshold': growth_threshold,
        'new_open_threshold': new_open_threshold,
    }


def calculate_risk_score(row: pd.Series) -> int:
    """Estimate a risk/churn score for a customer row."""
    if row['TOTAL_T1'] == 0:
        base = 30 if row['BIEN_DONG'] == 'MO_MOI' else 50
        return min(100, int(base))

    severity = abs(row['DELTA']) / max(row['TOTAL_T1'], 1)
    score = severity * 100
    if row['BIEN_DONG'] == 'TAT_TOAN':
        score += 30
    elif row['BIEN_DONG'] == 'GIAM':
        score += 20
    elif row['BIEN_DONG'] == 'MO_MOI':
        score += 10
    return min(100, int(round(score)))


def build_segment_summary(comparison_df: pd.DataFrame) -> pd.DataFrame:
    """Build a summary by customer balance bucket."""
    if comparison_df.empty:
        return pd.DataFrame()

    df = comparison_df.copy()
    labels = [bucket[0] for bucket in SEGMENT_BUCKETS]
    bins = [bucket[1] for bucket in SEGMENT_BUCKETS] + [SEGMENT_BUCKETS[-1][2]]
    df['BALANCE_BUCKET'] = pd.cut(
        df['TOTAL_T2'],
        bins=bins,
        labels=labels,
        right=False,
        include_lowest=True,
    ).astype(str)

    grouped = df.groupby('BALANCE_BUCKET', as_index=False).agg({
        'MA_KH': 'count',
        'TOTAL_T1': 'sum',
        'TOTAL_T2': 'sum',
        'DELTA': 'sum',
        'CHURN_SCORE': 'mean',
    }).rename(columns={
        'MA_KH': 'SO_KH',
        'TOTAL_T1': 'TONG_T1',
        'TOTAL_T2': 'TONG_T2',
        'DELTA': 'TONG_DELTA',
        'CHURN_SCORE': 'DIEM_RUI_RO_TRUNG_BINH',
    })

    grouped['TY_LE_TANG_TRUONG'] = grouped.apply(
        lambda r: (r['TONG_DELTA'] / r['TONG_T1'] * 100) if r['TONG_T1'] != 0 else 0,
        axis=1,
    )

    change_counts = df.groupby(['BALANCE_BUCKET', 'BIEN_DONG']).size().unstack(fill_value=0)
    for status in ['MO_MOI', 'TAT_TOAN', 'TANG', 'GIAM', 'KHONG_DOI']:
        if status in change_counts.columns:
            grouped[status] = change_counts[status].values
        else:
            grouped[status] = 0

    return grouped[['BALANCE_BUCKET', 'SO_KH', 'TONG_T1', 'TONG_T2', 'TONG_DELTA', 'TY_LE_TANG_TRUONG', 'MO_MOI', 'TAT_TOAN', 'TANG', 'GIAM', 'KHONG_DOI', 'DIEM_RUI_RO_TRUNG_BINH']]


def build_alert_summary(
    summary_branch: pd.DataFrame,
    summary_change: pd.DataFrame,
    outliers_df: pd.DataFrame,
    segment_summary: pd.DataFrame,
    recommendations_df: pd.DataFrame,
) -> pd.DataFrame:
    """Build a list of alerts for business review."""
    rows = []

    if not summary_branch.empty:
        for row in summary_branch.itertuples(index=False):
            if row.TONG_DELTA < 0 and row.TONG_DELTA <= -ALERT_BRANCH_DELTA_THRESHOLD:
                rows.append({
                    'LOAI_ALERT': 'Chi nhánh giảm mạnh',
                    'MA_CN': row.MA_CN,
                    'GIA_TRI': f"{row.TONG_DELTA:,.0f}",
                    'DIEN_GIAI': f"Chi nhánh {row.MA_CN} giảm {row.TONG_DELTA:,.0f} so với quý trước.",
                })
            elif row.TY_LE_TANG_TRUONG <= ALERT_BRANCH_DROP_RATE:
                rows.append({
                    'LOAI_ALERT': 'Tốc độ tăng trưởng thấp',
                    'MA_CN': row.MA_CN,
                    'GIA_TRI': f"{row.TY_LE_TANG_TRUONG:.2%}",
                    'DIEN_GIAI': f"Chi nhánh {row.MA_CN} có tỷ lệ tăng trưởng {row.TY_LE_TANG_TRUONG:.2%}.",
                })

    if not summary_change.empty:
        weak_segments = summary_change[summary_change['TONG_DELTA'] < 0]
        for row in weak_segments.itertuples(index=False):
            if abs(row.TONG_DELTA) >= ALERT_BRANCH_DELTA_THRESHOLD:
                rows.append({
                    'LOAI_ALERT': 'Nhóm biến động giảm lớn',
                    'MA_CN': '',
                    'GIA_TRI': f"{row.TONG_DELTA:,.0f}",
                    'DIEN_GIAI': f"Nhóm biến động {row.BIEN_DONG} giảm lớn với tổng DELTA {row.TONG_DELTA:,.0f}.",
                })

    if not outliers_df.empty:
        rows.append({
            'LOAI_ALERT': 'Phát hiện bất thường',
            'MA_CN': '',
            'GIA_TRI': len(outliers_df),
            'DIEN_GIAI': f"Có {len(outliers_df)} khách hàng bị phát hiện bất thường theo IQR.",
        })

    if not recommendations_df.empty:
        high_risk = len(recommendations_df[recommendations_df['MUC_UU_TIEN'] == 'RAT_CAO'])
        if high_risk > 0:
            rows.append({
                'LOAI_ALERT': 'Rủi ro khách hàng cao',
                'MA_CN': '',
                'GIA_TRI': high_risk,
                'DIEN_GIAI': f"Có {high_risk} khách hàng ở mức ưu tiên RẤT CAO cần giải pháp giữ chân.",
            })

    if not segment_summary.empty:
        negative_bucket = segment_summary[segment_summary['TONG_DELTA'] < 0]
        if not negative_bucket.empty:
            worst_bucket = negative_bucket.sort_values('TONG_DELTA').iloc[0]
            rows.append({
                'LOAI_ALERT': 'Phân khúc giảm',
                'MA_CN': worst_bucket.BALANCE_BUCKET,
                'GIA_TRI': f"{worst_bucket.TONG_DELTA:,.0f}",
                'DIEN_GIAI': f"Phân khúc {worst_bucket.BALANCE_BUCKET} giảm {worst_bucket.TONG_DELTA:,.0f}.",
            })

    if not rows:
        return pd.DataFrame()

    alert_df = pd.DataFrame(rows)
    alert_df.index = range(1, len(alert_df) + 1)
    return alert_df[['LOAI_ALERT', 'MA_CN', 'GIA_TRI', 'DIEN_GIAI']]


def build_top_customers(comparison_df: pd.DataFrame, limit: int = 20) -> dict:
    """Build top gainers, top losers and top value customer tables."""
    if comparison_df.empty:
        return {
            'top_gainers': pd.DataFrame(),
            'top_losers': pd.DataFrame(),
            'top_value': pd.DataFrame(),
        }

    top_gainers = comparison_df.sort_values('DELTA', ascending=False).head(limit).copy()
    top_losers = comparison_df.sort_values('DELTA', ascending=True).head(limit).copy()
    top_value = comparison_df.sort_values('TOTAL_T2', ascending=False).head(limit).copy()

    for df in [top_gainers, top_losers, top_value]:
        for col in ['TOTAL_T1', 'TOTAL_T2', 'DELTA']:
            if col in df.columns:
                df[col] = df[col].astype(float)

    return {
        'top_gainers': top_gainers,
        'top_losers': top_losers,
        'top_value': top_value,
    }


def get_last_run_metrics() -> dict:
    """Read the last two history records and return summary comparison."""
    history_df = load_history(limit=2)
    if len(history_df) < 2:
        return {}

    current = history_df.iloc[0]
    previous = history_df.iloc[1]
    return {
        'current': current,
        'previous': previous,
        'delta_customers': int(current['customers'] - previous['customers']) if 'customers' in current and 'customers' in previous else None,
        'delta_total': float(current['total_delta'] - previous['total_delta']) if 'total_delta' in current and 'total_delta' in previous else None,
    }


def build_change_type_trends(comparison_df: pd.DataFrame) -> pd.DataFrame:
    """Build trend counts by change type for dashboard charts."""
    if comparison_df.empty:
        return pd.DataFrame()

    trends = comparison_df.groupby('BIEN_DONG', as_index=False).agg({
        'MA_KH': 'count',
        'DELTA': 'sum'
    }).rename(columns={
        'MA_KH': 'SO_KH',
        'DELTA': 'TONG_DELTA'
    })
    return trends.sort_values('SO_KH', ascending=False)


def format_money(value) -> str:
    if pd.isna(value):
        return '0'
    return f"{value:,.0f}"


def build_driver_table(summary_df: pd.DataFrame, group_col: str, analysis_group: str) -> pd.DataFrame:
    """Build driver contribution table from summary data."""
    if summary_df.empty or group_col not in summary_df.columns:
        return pd.DataFrame()

    driver_df = summary_df[[group_col, 'TONG_DELTA']].copy()
    driver_df = driver_df.rename(columns={group_col: 'NHOM'})

    net_total = float(driver_df['TONG_DELTA'].sum())
    abs_total = float(driver_df['TONG_DELTA'].abs().sum())

    if net_total != 0:
        driver_df['DONG_GOP_RONG_%'] = driver_df['TONG_DELTA'] / net_total * 100
    else:
        driver_df['DONG_GOP_RONG_%'] = 0.0

    if abs_total != 0:
        driver_df['DONG_GOP_ABS_%'] = driver_df['TONG_DELTA'].abs() / abs_total * 100
    else:
        driver_df['DONG_GOP_ABS_%'] = 0.0

    driver_df['XU_HUONG'] = driver_df['TONG_DELTA'].apply(
        lambda value: 'TANG' if value > 0 else ('GIAM' if value < 0 else 'TRUNG_TINH')
    )
    driver_df['NHOM_PHAN_TICH'] = analysis_group
    driver_df['ABS_DELTA'] = driver_df['TONG_DELTA'].abs()

    return driver_df[[
        'NHOM_PHAN_TICH',
        'NHOM',
        'TONG_DELTA',
        'DONG_GOP_RONG_%',
        'DONG_GOP_ABS_%',
        'XU_HUONG',
        'ABS_DELTA',
    ]].sort_values('ABS_DELTA', ascending=False).reset_index(drop=True)


def main():
    """Main application function."""
    init_session_state()

    # Header
    st.markdown('<div class="main-header">📊 Hệ Thống So Sánh Dữ Liệu Tiền Gửi</div>', unsafe_allow_html=True)
    st.markdown('<div class="sub-header">So sánh dữ liệu tiền gửi giữa 2 quý (T1 và T2)</div>', unsafe_allow_html=True)

    # Sidebar wizard
    with st.sidebar:
        st.header("🧭 Hướng dẫn 1-2-3")
        st.write('Upload dữ liệu → Validate → So sánh → Xuất báo cáo')

        with st.expander('🔧 Mapping Cột', expanded=False):
            st.caption('Điều chỉnh tên cột nguồn về chuẩn trước khi chạy validate.')
            if st.button('↩️ Reset về chuẩn', use_container_width=True):
                for canonical in CANONICAL_COLUMNS:
                    st.session_state[f'map_{canonical}'] = canonical
                st.rerun()

            for canonical in CANONICAL_COLUMNS:
                key = f'map_{canonical}'
                if key not in st.session_state:
                    st.session_state[key] = canonical
                st.text_input(f'{canonical}', key=key)

            current_mapping = get_current_column_mapping()
            st.download_button(
                label='⬇️ Tải template CSV chuẩn',
                data=generate_template_csv(current_mapping),
                file_name='template_so_sanh_tien_gui.csv',
                mime='text/csv',
                use_container_width=True,
                help='Template dùng đúng mapping cột hiện tại.',
            )

        with st.expander('📁 Upload dữ liệu', expanded=True):
            t1_files = st.file_uploader(
                'T1 - Quý trước',
                type='csv',
                accept_multiple_files=True,
                key='t1_files',
            )
            t2_files = st.file_uploader(
                'T2 - Quý đối chiếu',
                type='csv',
                accept_multiple_files=True,
                key='t2_files',
            )

            if t1_files:
                st.info(f'Thêm {len(t1_files)} file T1')
            if t2_files:
                st.info(f'Thêm {len(t2_files)} file T2')

        current_mapping = get_current_column_mapping()
        has_upload = bool(st.session_state.get('t1_files') and st.session_state.get('t2_files'))
        has_validation = st.session_state.validation_result is not None
        has_processed = st.session_state.processed_data is not None

        st.markdown('---')
        st.subheader('🚦 Trạng thái')
        st.write(
            f"- {'✅' if has_upload else '⬜'} Upload dữ liệu\n"
            f"- {'✅' if has_validation else '⬜'} Validate\n"
            f"- {'✅' if has_processed else '⬜'} So sánh\n"
            f"- {'✅' if has_processed else '⬜'} Xuất báo cáo"
        )

        validate_clicked = st.button('2️⃣ Kiểm tra dữ liệu', disabled=not has_upload, use_container_width=True)
        compare_clicked = st.button('3️⃣ So sánh dữ liệu', disabled=not has_validation, type='primary', use_container_width=True)

        if validate_clicked:
            try:
                temp_dir = tempfile.mkdtemp()
                t1_paths = save_uploaded_files(t1_files, temp_dir)
                t2_paths = save_uploaded_files(t2_files, temp_dir)
                validation = validate_all_files(t1_paths, t2_paths, column_mapping=current_mapping)

                t1_meta = collect_file_metadata(t1_files)
                t2_meta = collect_file_metadata(t2_files)
                cache_key = build_cache_key(t1_meta, t2_meta, current_mapping)

                st.session_state.validation_result = validation
                st.session_state.upload_context = {
                    'temp_dir': temp_dir,
                    't1_paths': t1_paths,
                    't2_paths': t2_paths,
                    't1_meta': t1_meta,
                    't2_meta': t2_meta,
                    'cache_key': cache_key,
                    'column_mapping': current_mapping,
                }
                st.session_state.processed_data = None
                st.success('✅ Dữ liệu hợp lệ. Có thể chuyển sang bước 3.')
            except ValidationError as e:
                st.session_state.validation_result = None
                st.session_state.upload_context = None
                render_friendly_error(str(e))
            except Exception as e:
                st.session_state.validation_result = None
                st.session_state.upload_context = None
                render_friendly_error(f'Lỗi không xác định: {str(e)}')

        if compare_clicked:
            try:
                upload_context = st.session_state.upload_context
                if not upload_context:
                    st.warning('Vui lòng chạy bước 2 trước.')
                elif upload_context['column_mapping'] != current_mapping:
                    st.warning('Bạn đã đổi mapping cột. Vui lòng validate lại ở bước 2.')
                else:
                    progress = st.progress(0)
                    status = st.empty()

                    cache_key = upload_context['cache_key']
                    cache_hit = cache_key in st.session_state.processing_cache

                    status.text('⚙️ Đang xử lý đối chiếu...')
                    progress.progress(35)

                    if cache_hit:
                        processed = st.session_state.processing_cache[cache_key]
                    else:
                        processed = process_data(
                            upload_context['t1_paths'],
                            upload_context['t2_paths'],
                            st.session_state.validation_result,
                            column_mapping=current_mapping,
                        )
                        st.session_state.processing_cache[cache_key] = processed

                    status.text('💾 Đang lưu kết quả...')
                    progress.progress(85)
                    st.session_state.processed_data = processed
                    progress.progress(100)
                    status.text('✅ Hoàn tất')

                    comparison_df = processed['comparison']
                    append_history({
                        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        'cache_key': cache_key,
                        't1_input_files': ';'.join([item['name'] for item in upload_context['t1_meta']]),
                        't2_input_files': ';'.join([item['name'] for item in upload_context['t2_meta']]),
                        't1_files': len(upload_context['t1_meta']),
                        't2_files': len(upload_context['t2_meta']),
                        't1_date': st.session_state.validation_result.get('t1_date', ''),
                        't2_date': st.session_state.validation_result.get('t2_date', ''),
                        'branches': len(st.session_state.validation_result.get('branch_dates', {})),
                        'customers': len(comparison_df),
                        'total_t1': round(float(comparison_df['TOTAL_T1'].sum()), 2) if not comparison_df.empty else 0,
                        'total_t2': round(float(comparison_df['TOTAL_T2'].sum()), 2) if not comparison_df.empty else 0,
                        'total_delta': round(float(comparison_df['DELTA'].sum()), 2) if not comparison_df.empty else 0,
                        'high_risk_customers': len(processed['recommendations'][processed['recommendations']['MUC_UU_TIEN'] == 'RAT_CAO']) if not processed['recommendations'].empty else 0,
                        'cache_hit': 'YES' if cache_hit else 'NO',
                    })

                    # Save processed data to persistent cache
                    if not cache_hit:
                        save_processed_data(cache_key, processed)

                    st.success('✅ So sánh thành công!')
            except Exception as e:
                render_friendly_error(f'Lỗi xử lý bước 3: {str(e)}')

        with st.expander('🕘 Lịch sử so sánh', expanded=False):
            history_df = load_history()
            if history_df.empty:
                st.caption('Chưa có lịch sử xử lý.')
            else:
                # Add history selection for loading full results (only if cache_key column exists)
                if 'cache_key' in history_df.columns:
                    history_options = []
                    for idx, row in history_df.iterrows():
                        option = f"{row['timestamp']} - {row['t1_date']}→{row['t2_date']} ({row['customers']} KH, {row.get('total_delta', 0):,.0f}đ)"
                        history_options.append((row['cache_key'], option))

                    if history_options:  # Only show selectbox if there are options
                        # Create mapping for format_func to avoid StopIteration
                        cache_to_option = {opt[0]: opt[1] for opt in history_options}
                        
                        selected_history = st.selectbox(
                            'Chọn lịch sử để xem chi tiết:',
                            options=[opt[0] for opt in history_options],
                            format_func=lambda x: cache_to_option.get(x, f"Cache {x}"),
                            key='history_selector'
                        )

                        if st.button('📊 Tải kết quả phân tích', use_container_width=True):
                            with st.spinner('Đang tải kết quả lịch sử...'):
                                loaded_data = load_processed_data(selected_history)
                                if loaded_data and 'comparison' in loaded_data and not loaded_data['comparison'].empty:
                                    st.session_state.history_loaded_data = loaded_data
                                    st.session_state.history_timestamp = next((getattr(row, 'timestamp', 'Unknown') for row in history_df.itertuples() if getattr(row, 'cache_key', None) == selected_history), "Unknown")
                                    st.success('✅ Đã tải kết quả lịch sử! Xem ở tab "📊 Phân tích"')
                                    st.rerun()
                                else:
                                    st.error('❌ Không thể tải kết quả lịch sử này (cache không khả dụng)')
                    else:
                        st.caption('Chưa có lịch sử nào hỗ trợ xem chi tiết.')
                else:
                    st.info('💡 Tính năng xem chi tiết lịch sử sẽ khả dụng cho các lần chạy mới.')

                st.divider()
                col1, col2 = st.columns(2)
                with col1:
                    st.download_button(
                        label='⬇️ Tải lịch sử CSV',
                        data=history_df.to_csv(index=False).encode('utf-8-sig'),
                        file_name='comparison_history_recent.csv',
                        mime='text/csv',
                        use_container_width=True,
                    )
                with col2:
                    if st.button('🧹 Dọn dẹp cache cũ', use_container_width=True, help='Xóa cache entries cũ hơn 30 ngày'):
                        cleanup_old_cache(30)

        with st.expander('💡 Mẹo nhanh'):
            st.markdown(
                """
                - Tên file: `{MA_CN}_dp01_yyyymmdd.csv`
                - Nếu cột nguồn khác chuẩn, chỉ cần đổi ở phần Mapping Cột
                - Upload lại cùng file sẽ tận dụng cache để chạy nhanh hơn
                - Dùng template để gửi format chuẩn cho các đơn vị
                """
            )

    # Main content area
    if 'history_loaded_data' in st.session_state and st.session_state.history_loaded_data:
        st.info(f"📚 Đang xem kết quả lịch sử: {st.session_state.get('history_timestamp', 'N/A')}")
        if st.button('🔄 Quay lại kết quả hiện tại', key='back_to_current'):
            del st.session_state.history_loaded_data
            if 'history_timestamp' in st.session_state:
                del st.session_state.history_timestamp
            st.rerun()
        st.divider()
        data = st.session_state.history_loaded_data
    elif st.session_state.processed_data is not None:
        data = st.session_state.processed_data
    else:
        data = None

    if data is not None:
        comparison_df = data['comparison']

        st.markdown('### 📊 Tóm Tắt Nhanh')
        branch_count = len(data['summary_branch'])
        new_customers = int((comparison_df['BIEN_DONG'] == 'MO_MOI').sum())
        churn_customers = int((comparison_df['BIEN_DONG'] == 'TAT_TOAN').sum())
        growth_customers = int((comparison_df['BIEN_DONG'] == 'TANG').sum())
        total_delta = float(comparison_df['DELTA'].sum())
        average_delta = float(comparison_df['DELTA'].mean()) if not comparison_df.empty else 0
        alert_count = len(data['alert_summary']) if data.get('alert_summary') is not None else 0

        col1, col2, col3, col4, col5 = st.columns(5)
        with col1:
            st.metric('🏢 Chi Nhánh', branch_count)
        with col2:
            st.metric('📈 Tổng KH', len(comparison_df))
        with col3:
            st.metric('🆕 Mở Mới', new_customers)
        with col4:
            st.metric('🔚 Tất Toán', churn_customers)
        with col5:
            st.metric('💰 Tổng Δ', f"{total_delta:,.0f}")

        subcol1, subcol2, subcol3, subcol4 = st.columns(4)
        with subcol1:
            st.metric('📊 Tăng', growth_customers)
        with subcol2:
            st.metric('⚠️ Cảnh báo', alert_count)
        with subcol3:
            st.metric('📉 Trung bình Δ', f"{average_delta:,.0f}")
        with subcol4:
            st.metric('📈 Tỷ lệ tăng trưởng', f"{(total_delta / max(float(comparison_df['TOTAL_T1'].sum()), 1) * 100):.2f}%")

        st.divider()
        st.subheader('4️⃣ Chọn báo cáo để xem / xuất')
        tabs = st.tabs(['📊 Tổng quan', '📋 Chi tiết', '⚠️ Cảnh báo', '🎯 Khuyến nghị', '💾 Báo cáo'])

        with tabs[0]:
            tab_dashboard(data)

        with tabs[1]:
            tab_customer_details(data)
            st.divider()
            tab_branch_summary(data)
            st.divider()
            tab_customer_type_summary(data)
            st.divider()
            tab_segment_analysis(data)
            st.divider()
            tab_product_summary(data)
            st.divider()
            tab_top_customers(data)

        with tabs[2]:
            tab_alerts(data)
            st.divider()
            tab_outliers(data)

        with tabs[3]:
            tab_action_recommendations(data)
            st.divider()
            tab_driver_analysis(data)

        with tabs[4]:
            tab_charts(data)
            st.divider()
            tab_export(data)
    else:
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            st.markdown(
                """
                <div class="step-guide">
                    <h3>👋 Quy Trình 4 Bước</h3>
                    <ol>
                        <li>Upload dữ liệu T1 và T2</li>
                        <li>Validate cấu trúc và tên file</li>
                        <li>Chạy so sánh</li>
                        <li>Xem báo cáo và xuất Excel</li>
                    </ol>
                    <p>👉 Nếu dữ liệu nguồn khác tên cột chuẩn, chỉnh trong phần <b>Mapping Cột</b>.</p>
                </div>
                """,
                unsafe_allow_html=True,
            )


def process_data(t1_paths, t2_paths, validation, column_mapping=None):
    """Process uploaded data and generate statistics."""
    
    # Create file mapping
    file_mapping = {}
    branch_dates = validation['branch_dates']
    
    # Create mapping from file paths
    t1_by_name = {os.path.basename(p): p for p in t1_paths}
    t2_by_name = {os.path.basename(p): p for p in t2_paths}
    
    for ma_cn in branch_dates.keys():
        t1_date, t2_date = branch_dates[ma_cn]
        t1_filename = f"{ma_cn}_dp01_{t1_date}.csv"
        t2_filename = f"{ma_cn}_dp01_{t2_date}.csv"
        
        if t1_filename in t1_by_name and t2_filename in t2_by_name:
            file_mapping[ma_cn] = (t1_by_name[t1_filename], t2_by_name[t2_filename])
    
    # Load and process data
    all_raw_data = {}
    all_aggregated = {}
    comparison_results = []
    
    for ma_cn, (t1_file, t2_file) in file_mapping.items():
        # Load data
        df_t1 = load_and_normalize_csv(t1_file, column_mapping=column_mapping)
        df_t2 = load_and_normalize_csv(t2_file, column_mapping=column_mapping)
        
        all_raw_data[ma_cn] = {'T1': df_t1, 'T2': df_t2}
        
        # Filter invalid data
        df_t1_filtered = filter_valid_data(df_t1)
        df_t2_filtered = filter_valid_data(df_t2)
        
        # Aggregate
        agg_t1, agg_t2 = aggregate_pair(df_t1_filtered, df_t2_filtered)
        all_aggregated[ma_cn] = {'T1': agg_t1, 'T2': agg_t2}
        
        # Compare
        comparison = merge_and_compare(agg_t1, agg_t2, ma_cn)
        comparison_results.append(comparison)
    
    # Combine all comparisons
    comparison_df = pd.concat(comparison_results, ignore_index=True) if comparison_results else pd.DataFrame()
    if not comparison_df.empty:
        comparison_df['CHURN_SCORE'] = comparison_df.apply(calculate_risk_score, axis=1)

    # Generate summaries
    summary_branch_data = summary_by_branch(comparison_df)
    summary_cust_type = summary_by_customer_type(comparison_df)
    summary_product = summary_by_product_group(all_raw_data)
    summary_change = summary_by_change_type(comparison_df)

    # Build analysis and recommendations
    recommendations_df, recommendation_thresholds = build_action_recommendations(comparison_df)
    driver_customer = build_driver_table(summary_cust_type, 'CUST_TYPE_NAME', 'PHAN_KHUC')
    driver_product = build_driver_table(summary_product, 'DP_GROUP', 'SAN_PHAM')
    driver_analysis = pd.concat([driver_customer, driver_product], ignore_index=True)

    segment_summary = build_segment_summary(comparison_df)
    outliers = detect_outliers_iqr(comparison_df, column='DELTA')
    alert_summary = build_alert_summary(summary_branch_data, summary_change, outliers, segment_summary, recommendations_df)
    trend_summary = build_change_type_trends(comparison_df)
    top_customers = build_top_customers(comparison_df, limit=20)

    return {
        'comparison': comparison_df,
        'summary_branch': summary_branch_data,
        'summary_cust_type': summary_cust_type,
        'summary_product': summary_product,
        'summary_change': summary_change,
        'recommendations': recommendations_df,
        'recommendation_thresholds': recommendation_thresholds,
        'driver_customer': driver_customer,
        'driver_product': driver_product,
        'driver_analysis': driver_analysis,
        'segment_summary': segment_summary,
        'alert_summary': alert_summary,
        'trend_summary': trend_summary,
        'top_customers': top_customers,
        'outliers': outliers,
        'all_raw_data': all_raw_data,
        'validation': st.session_state.validation_result
    }


def tab_customer_details(data):
    """Display detailed customer information."""
    st.header("📋 Chi Tiết Khách Hàng")
    
    comparison_df = data['comparison']
    
    # Filters
    col1, col2, col3 = st.columns(3)
    
    with col1:
        selected_branches = st.multiselect(
            "Chi Nhánh",
            comparison_df['MA_CN'].unique(),
            default=comparison_df['MA_CN'].unique(),
            key="branch_filter"
        )
    
    with col2:
        selected_types = st.multiselect(
            "Phân Khúc",
            comparison_df['CUST_TYPE_NAME'].unique(),
            default=comparison_df['CUST_TYPE_NAME'].unique(),
            key="type_filter"
        )
    
    with col3:
        selected_changes = st.multiselect(
            "Biến Động",
            comparison_df['BIEN_DONG'].unique(),
            default=comparison_df['BIEN_DONG'].unique(),
            key="change_filter"
        )
    
    # Apply filters
    filtered_df = comparison_df[
        (comparison_df['MA_CN'].isin(selected_branches)) &
        (comparison_df['CUST_TYPE_NAME'].isin(selected_types)) &
        (comparison_df['BIEN_DONG'].isin(selected_changes))
    ].copy()
    
    # Format for display
    display_df = filtered_df.copy()
    for col in ['TOTAL_T1', 'TOTAL_T2', 'DELTA']:
        display_df[col] = display_df[col].apply(lambda x: f"{x:,.0f}")
    
    st.dataframe(display_df, use_container_width=True, height=400)
    
    # Statistics
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Tổng Khách Hàng", len(filtered_df))
    with col2:
        st.metric("Mở Mới", len(filtered_df[filtered_df['BIEN_DONG'] == 'MO_MOI']))
    with col3:
        st.metric("Tất Toán", len(filtered_df[filtered_df['BIEN_DONG'] == 'TAT_TOAN']))
    with col4:
        st.metric("Tăng", len(filtered_df[filtered_df['BIEN_DONG'] == 'TANG']))
    
    # Export filtered data to Excel
    if not filtered_df.empty:
        st.markdown("---")
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            if st.button("📥 Xuất Excel (Dữ Liệu Đã Lọc)", use_container_width=True, key="export_filtered"):
                try:
                    # Create fast Excel export
                    export_df = filtered_df.copy()
                    # Format numeric columns
                    for col in ['TOTAL_T1', 'TOTAL_T2', 'DELTA']:
                        if col in export_df.columns:
                            export_df[col] = export_df[col].apply(lambda x: f"{x:,.0f}" if pd.notna(x) else "0")
                    
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        export_df.to_excel(writer, sheet_name='Khách Hàng Đã Lọc', index=False)
                        
                        ws = writer.sheets['Khách Hàng Đã Lọc']
                        from openpyxl.styles import Font, PatternFill
                        from openpyxl.utils import get_column_letter
                        
                        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                        header_font = Font(bold=True, color="FFFFFF")
                        
                        for cell in ws[1]:
                            cell.fill = header_fill
                            cell.font = header_font
                        
                        for col_num in range(1, len(export_df.columns) + 1):
                            ws.column_dimensions[get_column_letter(col_num)].width = 15
                    
                    output.seek(0)
                    excel_data = output.getvalue()
                    
                    # Generate filename with timestamp
                    from datetime import datetime
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    filename = f"ChiTietKhachHang_Loc_{timestamp}.xlsx"
                    
                    st.download_button(
                        label="💾 Tải Xuống Excel",
                        data=excel_data,
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml"
                    )
                    
                    st.success(f"✅ Đã chuẩn bị file Excel: {filename}")
                    
                except Exception as e:
                    st.error(f"❌ Lỗi xuất Excel: {str(e)}")


def tab_dashboard(data):
    """Display the new dashboard view with KPI, alerts, and trends."""
    st.header("📈 Dashboard Tổng Quan")

    comparison_df = data['comparison']
    trend_summary = data.get('trend_summary', pd.DataFrame())
    alert_summary = data.get('alert_summary', pd.DataFrame())
    last_run = get_last_run_metrics()

    if comparison_df.empty:
        st.info("Chưa có dữ liệu để hiển thị dashboard.")
        return

    total_customers = len(comparison_df)
    total_delta = float(comparison_df['DELTA'].sum())
    total_t1 = float(comparison_df['TOTAL_T1'].sum())
    total_t2 = float(comparison_df['TOTAL_T2'].sum())
    rate_growth = (total_delta / total_t1 * 100) if total_t1 != 0 else 0
    new_customers = int((comparison_df['BIEN_DONG'] == 'MO_MOI').sum())
    churn_customers = int((comparison_df['BIEN_DONG'] == 'TAT_TOAN').sum())
    risk_customers = int((comparison_df['CHURN_SCORE'] >= 60).sum())

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.metric('Tổng khách hàng', total_customers)
    with c2:
        st.metric('Tổng T1', f"{total_t1:,.0f}")
    with c3:
        st.metric('Tổng T2', f"{total_t2:,.0f}")
    with c4:
        st.metric('Tăng trưởng', f"{rate_growth:.2f}%")

    c5, c6, c7, c8 = st.columns(4)
    with c5:
        st.metric('Mới', new_customers)
    with c6:
        st.metric('Tất toán', churn_customers)
    with c7:
        st.metric('Rủi ro cao', risk_customers)
    with c8:
        st.metric('Cảnh báo', len(alert_summary))

    if last_run:
        diff_customers = last_run.get('delta_customers')
        diff_total = last_run.get('delta_total')
        with st.expander('🕘 So sánh với lần chạy trước', expanded=True):
            st.write(f"Số lượng khách hiện tại: {total_customers} ({'+' if diff_customers >= 0 else ''}{diff_customers} so với lần trước)")
            st.write(f"Tổng DELTA hiện tại: {total_delta:,.0f} ({'+' if diff_total >= 0 else ''}{diff_total:,.0f} so với lần trước)")

    with st.expander('📌 Các cảnh báo chính', expanded=True):
        if alert_summary.empty:
            st.success('Không có cảnh báo cấp cao.')
        else:
            st.dataframe(alert_summary, use_container_width=True)

    if not trend_summary.empty:
        st.subheader('Xu hướng biến động theo loại')
        fig = px.bar(
            trend_summary,
            x='BIEN_DONG',
            y='SO_KH',
            color='TONG_DELTA',
            title='Phân bổ khách hàng theo nhóm biến động',
            color_continuous_scale=['red', 'yellow', 'green'],
        )
        st.plotly_chart(fig, use_container_width=True)

    st.subheader('Chi tiết nhóm biến động')
    st.dataframe(trend_summary, use_container_width=True)


def tab_segment_analysis(data):
    """Display advanced segment analysis by balance bucket."""
    st.header("🌐 Phân Khúc Nâng Cao")

    segment_summary = data.get('segment_summary', pd.DataFrame())
    if segment_summary.empty:
        st.info('Chưa có dữ liệu phân khúc để phân tích.')
        return

    display_df = segment_summary.copy()
    for col in ['TONG_T1', 'TONG_T2', 'TONG_DELTA', 'DIEM_RUI_RO_TRUNG_BINH', 'TY_LE_TANG_TRUONG']:
        if col in display_df.columns:
            display_df[col] = display_df[col].apply(lambda x: f"{x:,.0f}" if x is not None and col != 'TY_LE_TANG_TRUONG' else f"{x:.2f}%" if x is not None else '0')

    st.markdown(
        'Phân tích theo phân khúc dư tiền T2, cho biết phân khúc nào đóng góp tăng/giảm và mức độ rủi ro trung bình.'
    )
    st.dataframe(display_df, use_container_width=True)

    st.subheader('Biểu đồ phân khúc')
    fig = px.bar(
        segment_summary,
        x='BALANCE_BUCKET',
        y='TONG_DELTA',
        color='TONG_DELTA',
        title='Đóng góp DELTA theo phân khúc dư tiền',
        color_continuous_scale=['red', 'yellow', 'green'],
    )
    st.plotly_chart(fig, use_container_width=True)

    with st.expander('ℹ️ Gợi ý phân tích', expanded=False):
        st.markdown(
            """
            - Các phân khúc có DELTA âm lớn cần xem lại chính sách chăm sóc và sản phẩm phù hợp.
            - Phân khúc có điểm rủi ro trung bình cao cần ưu tiên giữ chân.
            - Dùng bảng này để xác định nhóm khách hàng mục tiêu cho chương trình marketing.
            """
        )


def tab_top_customers(data):
    """Display top customers by gain, loss and balance."""
    st.header("🏆 Khách Hàng Tiêu Biểu")

    top_customers = data.get('top_customers', {})
    top_gainers = top_customers.get('top_gainers', pd.DataFrame())
    top_losers = top_customers.get('top_losers', pd.DataFrame())
    top_value = top_customers.get('top_value', pd.DataFrame())

    if top_gainers.empty and top_losers.empty and top_value.empty:
        st.info('Chưa có dữ liệu khách hàng tiêu biểu.')
        return

    st.subheader('Top khách hàng tăng mạnh')
    if not top_gainers.empty:
        display_top = top_gainers.copy()
        for col in ['TOTAL_T1', 'TOTAL_T2', 'DELTA']:
            if col in display_top.columns:
                display_top[col] = display_top[col].apply(lambda x: f"{x:,.0f}")
        st.dataframe(display_top, use_container_width=True, height=260)

    st.subheader('Top khách hàng giảm mạnh')
    if not top_losers.empty:
        display_losers = top_losers.copy()
        for col in ['TOTAL_T1', 'TOTAL_T2', 'DELTA']:
            if col in display_losers.columns:
                display_losers[col] = display_losers[col].apply(lambda x: f"{x:,.0f}")
        st.dataframe(display_losers, use_container_width=True, height=260)

    st.subheader('Top khách hàng theo dư lớn nhất T2')
    if not top_value.empty:
        display_value = top_value.copy()
        for col in ['TOTAL_T1', 'TOTAL_T2', 'DELTA']:
            if col in display_value.columns:
                display_value[col] = display_value[col].apply(lambda x: f"{x:,.0f}")
        st.dataframe(display_value, use_container_width=True, height=260)


def tab_alerts(data):
    """Display alert summary and recommendations."""
    st.header("⚠️ Cảnh Báo & Nhận Diện Rủi Ro")

    alert_summary = data.get('alert_summary', pd.DataFrame())
    if alert_summary.empty:
        st.success('Không có cảnh báo cấp cao hiện tại.')
        return

    st.markdown(
        'Danh sách cảnh báo được xây dựng dựa trên chi nhánh giảm mạnh, nhóm biến động âm lớn, outlier và rủi ro khách hàng.'
    )
    st.dataframe(alert_summary, use_container_width=True, height=380)

    with st.expander('ℹ️ Cách đọc cảnh báo', expanded=False):
        st.markdown(
            """
            - `Chi nhánh giảm mạnh`: xem chi nhánh có DELTA âm lớn.
            - `Tốc độ tăng trưởng thấp`: cảnh báo chi nhánh có mức tăng trưởng dưới ngưỡng.
            - `Phát hiện bất thường`: số lượng khách bất thường theo IQR.
            - `Rủi ro khách hàng cao`: số khách hàng ưu tiên giữ chân ngay.
            """
        )


def tab_branch_summary(data):
    """Display summary by branch."""
    st.header("📊 Thống Kê Theo Chi Nhánh")
    
    st.info("📌 So sánh dữ liệu tiền gửi giữa T1 và T2 theo từng chi nhánh, bao gồm tỷ lệ tăng trưởng")
    
    summary_branch = data['summary_branch']
    
    # Format for display
    display_df = summary_branch.copy()
    for col in ['TONG_T1', 'TONG_T2', 'TONG_DELTA']:
        display_df[col] = display_df[col].apply(lambda x: f"{x:,.0f}")
    display_df['TY_LE_TANG_TRUONG'] = display_df['TY_LE_TANG_TRUONG'].apply(lambda x: f"{x:.2%}")
    
    st.dataframe(display_df, use_container_width=True)
    
    # Summary metrics
    col1, col2, col3 = st.columns(3)
    with col1:
        total_t1 = summary_branch['TONG_T1'].sum()
        st.metric("Tổng T1", f"{total_t1:,.0f}")
    with col2:
        total_t2 = summary_branch['TONG_T2'].sum()
        st.metric("Tổng T2", f"{total_t2:,.0f}")
    with col3:
        total_delta = summary_branch['TONG_DELTA'].sum()
        st.metric("Tổng DELTA", f"{total_delta:,.0f}")
    
    with st.expander("ℹ️ Giải Thích"):
        st.markdown("""
        **TONG_T1:** Tổng dư tiền của chi nhánh ở quý 1
        **TONG_T2:** Tổng dư tiền của chi nhánh ở quý 2
        **TONG_DELTA:** Chênh lệch (TONG_T2 - TONG_T1)
        **TY_LE_TANG_TRUONG:** Tỷ lệ tăng (hay giảm) so với quý trước
        """)


def tab_customer_type_summary(data):
    """Display summary by customer type."""
    st.header("👥 Thống Kê Theo Phân Khúc")
    
    st.info("📌 Phân tích dữ liệu giữa khách hàng Cá Nhân và Pháp Nhân")
    
    summary_cust = data['summary_cust_type']
    
    # Format for display
    display_df = summary_cust.copy()
    for col in ['TONG_T1', 'TONG_T2', 'TONG_DELTA']:
        display_df[col] = display_df[col].apply(lambda x: f"{x:,.0f}")
    
    st.dataframe(display_df, use_container_width=True)
    
    # Chart
    if not summary_cust.empty:
        fig = px.bar(
            summary_cust,
            x='CUST_TYPE_NAME',
            y=['TONG_T1', 'TONG_T2'],
            barmode='group',
            title='So Sánh Giữa T1 và T2 Theo Phân Khúc'
        )
        st.plotly_chart(fig, use_container_width=True)


def tab_product_summary(data):
    """Display summary by product group."""
    st.header("📦 Thống Kê Theo Nhóm Sản Phẩm")
    
    st.info("📌 Phân tích dữ liệu tiền gửi theo từng loại sản phẩm")
    
    summary_product = data['summary_product']
    
    # Format for display
    display_df = summary_product.copy()
    for col in ['TONG_T1', 'TONG_T2', 'TONG_DELTA']:
        display_df[col] = display_df[col].apply(lambda x: f"{x:,.0f}")
    
    st.dataframe(display_df, use_container_width=True)
    
    # Chart
    if not summary_product.empty:
        fig = px.bar(
            summary_product,
            x='DP_GROUP',
            y='TONG_DELTA',
            title='Biến Động Theo Nhóm Sản Phẩm',
            color='TONG_DELTA',
            color_continuous_scale=['red', 'yellow', 'green']
        )
        st.plotly_chart(fig, use_container_width=True)


def tab_outliers(data):
    """Display outliers."""
    st.header("🚨 Phát Hiện Bất Thường")
    
    st.warning("""
    ⚠️ **Bất Thường (Outlier)** là những giao dịch có chênh lệch bất thường so với trung bình.
    Hệ thống sử dụng phương pháp **IQR (Interquartile Range)** để phát hiện.
    """)
    
    outliers = data['outliers']
    
    if outliers.empty:
        st.success("✅ Không phát hiện bất thường nào")
    else:
        # Display outliers
        display_df = outliers.copy()
        for col in ['TOTAL_T1', 'TOTAL_T2', 'DELTA']:
            if col in display_df.columns:
                display_df[col] = display_df[col].apply(lambda x: f"{x:,.0f}")
        
        cols_to_show = ['MA_CN', 'MA_KH', 'TEN_KH', 'CUST_TYPE_NAME', 'TOTAL_T1', 'TOTAL_T2', 'DELTA', 'BIEN_DONG']
        display_df = display_df[[col for col in cols_to_show if col in display_df.columns]]
        
        st.dataframe(display_df, use_container_width=True, height=400)
        
        # Statistics
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Tổng Bất Thường", len(outliers))
        with col2:
            positive = len(outliers[outliers['DELTA'] > 0])
            st.metric("Tăng Bất Thường", positive)
        with col3:
            negative = len(outliers[outliers['DELTA'] < 0])
            st.metric("Giảm Bất Thường", negative)


def tab_action_recommendations(data):
    """Display action recommendations for customer follow-up."""
    st.header("🎯 Khuyến Nghị Hành Động")

    recommendations_df = data.get('recommendations', pd.DataFrame())
    thresholds = data.get('recommendation_thresholds', {})

    if recommendations_df.empty:
        st.info("Chưa phát sinh khuyến nghị hành động theo rule hiện tại.")
        return

    st.info(
        "📌 Hệ thống tự xếp hạng theo mức ưu tiên và đưa ra hành động gợi ý dựa trên biến động tiền gửi."
    )

    with st.expander("ℹ️ Giải thích logic khuyến nghị", expanded=False):
        st.markdown(
            """
            **1) Cách hệ thống xác định ngưỡng**
            - `Ngưỡng giảm mạnh`: max(50 triệu, Q75 của |DELTA âm|)
            - `Ngưỡng tăng mạnh`: max(50 triệu, Q75 của DELTA dương)
            - `Ngưỡng mở mới lớn`: max(100 triệu, Q75 của TOTAL_T2 với khách `MO_MOI`)

            **2) Quy tắc xếp mức ưu tiên**
            - `RAT_CAO`: Khách tất toán hoặc giảm rất mạnh
            - `CAO`: Khách giảm mạnh vượt ngưỡng
            - `TRUNG_BINH`: Khách mở mới lớn hoặc tăng mạnh

            **3) Ý nghĩa cột trong bảng**
            - `LY_DO`: Lý do rule được kích hoạt
            - `HANH_DONG_GOI_Y`: Gợi ý thao tác thực tế cho đội chăm sóc/kinh doanh
            - `ABS_DELTA`: Độ lớn biến động tuyệt đối để ưu tiên xử lý
            """
        )

    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Ngưỡng giảm mạnh", f"{thresholds.get('risk_threshold', 0):,.0f}")
    with col2:
        st.metric("Ngưỡng tăng mạnh", f"{thresholds.get('growth_threshold', 0):,.0f}")
    with col3:
        st.metric("Ngưỡng mở mới lớn", f"{thresholds.get('new_open_threshold', 0):,.0f}")

    colf1, colf2 = st.columns(2)
    with colf1:
        selected_priorities = st.multiselect(
            "Mức ưu tiên",
            recommendations_df['MUC_UU_TIEN'].unique(),
            default=list(recommendations_df['MUC_UU_TIEN'].unique()),
            key='recommendation_priority_filter',
        )
    with colf2:
        selected_branches = st.multiselect(
            "Chi nhánh",
            recommendations_df['MA_CN'].unique(),
            default=list(recommendations_df['MA_CN'].unique()),
            key='recommendation_branch_filter',
        )

    filtered_df = recommendations_df[
        recommendations_df['MUC_UU_TIEN'].isin(selected_priorities)
        & recommendations_df['MA_CN'].isin(selected_branches)
    ].copy()

    display_df = filtered_df.copy()
    for col in ['TOTAL_T1', 'TOTAL_T2', 'DELTA', 'ABS_DELTA', 'DIEM_RUI_RO']:
        if col in display_df.columns:
            display_df[col] = display_df[col].apply(lambda value: f"{value:,.0f}") if col != 'DIEM_RUI_RO' else display_df[col].apply(lambda value: int(value) if pd.notna(value) else 0)

    st.dataframe(display_df, use_container_width=True, height=420)

    m1, m2, m3 = st.columns(3)
    with m1:
        st.metric("Số KH cần hành động", len(filtered_df))
    with m2:
        high_count = len(filtered_df[filtered_df['MUC_UU_TIEN'] == 'RAT_CAO'])
        st.metric("Ưu tiên rất cao", high_count)
    with m3:
        risk_value = abs(filtered_df.loc[filtered_df['DELTA'] < 0, 'DELTA'].sum())
        st.metric("Giá trị rủi ro", f"{risk_value:,.0f}")


def tab_driver_analysis(data):
    """Display driver analysis by customer segment and product group."""
    st.header("🧭 Phân Tích Động Lực")

    driver_customer = data.get('driver_customer', pd.DataFrame())
    driver_product = data.get('driver_product', pd.DataFrame())

    if driver_customer.empty and driver_product.empty:
        st.info("Không có dữ liệu để phân tích động lực biến động.")
        return

    st.info(
        "📌 Cho biết nhóm nào đang đóng góp chính vào tăng/giảm tổng DELTA, theo phân khúc khách hàng và sản phẩm."
    )

    with st.expander("ℹ️ Giải thích cách đọc phân tích động lực", expanded=False):
        st.markdown(
            """
            **1) Các chỉ số chính**
            - `TONG_DELTA`: Mức đóng góp ròng của từng nhóm
            - `DONG_GOP_RONG_%`: Tỷ trọng đóng góp trên tổng DELTA ròng
            - `DONG_GOP_ABS_%`: Tỷ trọng theo trị tuyệt đối (cho biết mức ảnh hưởng thật sự)
            - `XU_HUONG`: `TANG`, `GIAM`, hoặc `TRUNG_TINH`

            **2) Cách diễn giải nhanh**
            - Nhóm có `TONG_DELTA` dương lớn: động lực tăng trưởng chính
            - Nhóm có `TONG_DELTA` âm lớn (về độ lớn): nguyên nhân kéo giảm chính
            - Ưu tiên nhóm có `DONG_GOP_ABS_%` cao để tác động hiệu quả hơn

            **3) Lưu ý nghiệp vụ**
            - `DONG_GOP_RONG_%` có thể vượt 100% khi các nhóm tăng và giảm triệt tiêu nhau
            - Vì vậy nên xem đồng thời cả `DONG_GOP_RONG_%` và `DONG_GOP_ABS_%`
            """
        )

    c1, c2 = st.columns(2)

    with c1:
        st.subheader("Theo phân khúc")
        if driver_customer.empty:
            st.caption("Không có dữ liệu phân khúc.")
        else:
            fig = px.bar(
                driver_customer.sort_values('TONG_DELTA', ascending=False),
                x='NHOM',
                y='TONG_DELTA',
                color='XU_HUONG',
                title='Đóng góp DELTA theo phân khúc',
            )
            st.plotly_chart(fig, use_container_width=True)

            display_cust = driver_customer.copy()
            display_cust['TONG_DELTA'] = display_cust['TONG_DELTA'].apply(lambda value: f"{value:,.0f}")
            display_cust['DONG_GOP_RONG_%'] = display_cust['DONG_GOP_RONG_%'].apply(lambda value: f"{value:.2f}%")
            display_cust['DONG_GOP_ABS_%'] = display_cust['DONG_GOP_ABS_%'].apply(lambda value: f"{value:.2f}%")
            st.dataframe(display_cust, use_container_width=True)

    with c2:
        st.subheader("Theo sản phẩm")
        if driver_product.empty:
            st.caption("Không có dữ liệu sản phẩm.")
        else:
            fig = px.bar(
                driver_product.sort_values('TONG_DELTA', ascending=False),
                x='NHOM',
                y='TONG_DELTA',
                color='XU_HUONG',
                title='Đóng góp DELTA theo sản phẩm',
            )
            st.plotly_chart(fig, use_container_width=True)

            display_product = driver_product.copy()
            display_product['TONG_DELTA'] = display_product['TONG_DELTA'].apply(lambda value: f"{value:,.0f}")
            display_product['DONG_GOP_RONG_%'] = display_product['DONG_GOP_RONG_%'].apply(lambda value: f"{value:.2f}%")
            display_product['DONG_GOP_ABS_%'] = display_product['DONG_GOP_ABS_%'].apply(lambda value: f"{value:.2f}%")
            st.dataframe(display_product, use_container_width=True)

    insight_col1, insight_col2 = st.columns(2)
    if not driver_customer.empty:
        top_pos_cust = driver_customer.sort_values('TONG_DELTA', ascending=False).iloc[0]
        top_neg_cust = driver_customer.sort_values('TONG_DELTA', ascending=True).iloc[0]
        with insight_col1:
            st.success(
                f"Phân khúc đóng góp tăng mạnh nhất: {top_pos_cust['NHOM']} ({top_pos_cust['TONG_DELTA']:,.0f})"
            )
            st.warning(
                f"Phân khúc kéo giảm mạnh nhất: {top_neg_cust['NHOM']} ({top_neg_cust['TONG_DELTA']:,.0f})"
            )

    if not driver_product.empty:
        top_pos_prod = driver_product.sort_values('TONG_DELTA', ascending=False).iloc[0]
        top_neg_prod = driver_product.sort_values('TONG_DELTA', ascending=True).iloc[0]
        with insight_col2:
            st.success(
                f"Sản phẩm đóng góp tăng mạnh nhất: {top_pos_prod['NHOM']} ({top_pos_prod['TONG_DELTA']:,.0f})"
            )
            st.warning(
                f"Sản phẩm kéo giảm mạnh nhất: {top_neg_prod['NHOM']} ({top_neg_prod['TONG_DELTA']:,.0f})"
            )


def tab_charts(data):
    """Display charts."""
    st.header("📈 Biểu Đồ")
    
    comparison_df = data['comparison']
    summary_branch = data['summary_branch']
    summary_cust = data['summary_cust_type']
    summary_product = data['summary_product']
    
    col1, col2 = st.columns(2)
    
    # Chart 1: Growth by branch
    with col1:
        if not summary_branch.empty:
            fig = px.bar(
                summary_branch.sort_values('TONG_DELTA', ascending=False),
                x='MA_CN',
                y='TONG_DELTA',
                title='Tăng Trưởng Theo Chi Nhánh',
                color='TONG_DELTA',
                color_continuous_scale=['red', 'yellow', 'green']
            )
            st.plotly_chart(fig, use_container_width=True)
    
    # Chart 2: Customer type comparison
    with col2:
        if not summary_cust.empty:
            fig = px.pie(
                summary_cust,
                names='CUST_TYPE_NAME',
                values='TONG_DELTA',
                title='Tỷ Lệ DELTA Theo Phân Khúc'
            )
            st.plotly_chart(fig, use_container_width=True)

    if not summary_branch.empty:
        st.subheader('Heatmap tăng trưởng chi nhánh')
        heatmap_matrix = [summary_branch['TY_LE_TANG_TRUONG'].tolist()]
        heatmap_labels = summary_branch['MA_CN'].tolist()
        fig = px.imshow(
            heatmap_matrix,
            x=heatmap_labels,
            y=['Tăng trưởng (%)'],
            color_continuous_scale='RdYlGn',
            labels={'x': 'Chi Nhánh', 'y': ''},
            text_auto='.2f',
        )
        fig.update_xaxes(tickangle=45)
        st.plotly_chart(fig, use_container_width=True)

    # Chart 3: Product group
    col3, col4 = st.columns(2)
    with col3:
        if not summary_product.empty:
            fig = px.bar(
                summary_product.sort_values('TONG_DELTA', ascending=False),
                x='DP_GROUP',
                y=['TONG_T1', 'TONG_T2'],
                barmode='group',
                title='So Sánh Theo Nhóm Sản Phẩm'
            )
            st.plotly_chart(fig, use_container_width=True)
    
    # Chart 4: Change type distribution
    with col4:
        change_summary = data['summary_change']
        if not change_summary.empty:
            fig = px.bar(
                change_summary,
                x='BIEN_DONG',
                y='SO_KH',
                title='Phân Bố Biến Động Khách Hàng',
                color='BIEN_DONG'
            )
            st.plotly_chart(fig, use_container_width=True)


def tab_export(data):
    """Export functionality."""
    st.header("💾 Xuất Báo Cáo Excel")
    
    st.info("""
    📌 **Tính Năng Xuất Báo Cáo:**
    - Xuất toàn bộ dữ liệu so sánh sang 7 sheet Excel
    - Định dạng chuyên nghiệp với header màu
    - Hỗ trợ công thức và định dạng số tiền
    """)
    
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        if st.button("📥 Tải Xuống Excel Báo Cáo Đầy Đủ", use_container_width=True, type="primary"):
            try:
                with st.spinner("⏳ Đang chuẩn bị file Excel..."):
                    excel_data = export_to_excel(
                        data['comparison'],
                        data['summary_branch'],
                        data['summary_cust_type'],
                        data['summary_product'],
                        data['outliers'],
                        data.get('recommendations'),
                        data.get('driver_analysis'),
                        data.get('segment_summary'),
                        data.get('alert_summary')
                    )
                    
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    filename = f"BaoCao_SoSanh_TienGui_{timestamp}.xlsx"
                    
                    st.download_button(
                        label="💾 Tải Xuống",
                        data=excel_data,
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                    
                    st.success(f"✅ File sẵn sàng: {filename}")
            except Exception as e:
                st.error(f"❌ Lỗi xuất báo cáo: {str(e)}")
    
    st.divider()
    
    # Sheet descriptions
    st.subheader("📋 Nội Dung Báo Cáo")
    
    tabs_info = st.tabs([
        "📊 Chi Tiết Khách Hàng",
        "🏢 Theo Chi Nhánh",
        "👥 Theo Phân Khúc",
        "🌐 Phân Khúc Nâng Cao",
        "📦 Theo Sản Phẩm",
        "⚠️ Bất Thường",
        "🎯 Khuyến Nghị",
        "🧭 Động Lực",
        "🚨 Cảnh Báo"
    ])
    
    with tabs_info[0]:
        st.markdown("""
        **Danh sách tất cả khách hàng với:**
        - Mã khách hàng, tên khách hàng
        - Dư tiền T1 và T2
        - Chênh lệch (DELTA)
        - Loại biến động
        
        **Dùng để:** Phân tích chi tiết từng giao dịch
        """)
    
    with tabs_info[1]:
        st.markdown("""
        **Thống kê tổng hợp theo chi nhánh:**
        - Tổng dư T1 và T2
        - Tổng chênh lệch
        - Tỷ lệ tăng trưởng (%)
        
        **Dùng để:** So sánh hiệu suất giữa các chi nhánh
        """)
    
    with tabs_info[2]:
        st.markdown("""
        **Thống kê theo phân khúc khách:**
        - Cá Nhân vs Pháp Nhân
        - Số lượng khách hàng
        - Tổng dư tiền theo phân khúc
        
        **Dùng để:** Phân tích hành vi từng khúc thị trường
        """)
    
    with tabs_info[3]:
        st.markdown("""
        **Phân khúc nâng cao theo dư tiền T2:**
        - Nhóm khách theo mức tiền gửi
        - Tổng dư tiền, DELTA, tỷ lệ tăng trưởng
        - Mức rủi ro trung bình của từng phân khúc

        **Dùng để:** Định hướng chương trình chăm sóc và sản phẩm
        """)
    
    with tabs_info[4]:
        st.markdown("""
        **Thống kê theo loại sản phẩm:**
        - Từng loại sản phẩm gửi
        - Tổng dư và chênh lệch
        - Số khách sử dụng

        **Dùng để:** Đánh giá hiệu suất sản phẩm
        """)

    with tabs_info[5]:
        st.markdown("""
        - Chi tiết khách hàng
        - Mức biến động
        
        **Dùng để:** Phát hiện giao dịch cần kiểm tra
        """)

    with tabs_info[6]:
        st.markdown("""
        **Danh sách khuyến nghị hành động:**
        - Mức ưu tiên (RAT_CAO/CAO/TRUNG_BINH)
        - Lý do kích hoạt khuyến nghị
        - Hành động gợi ý cho đội kinh doanh/chăm sóc
        - Bao gồm ngữ cảnh biến động (T1, T2, DELTA, BIEN_DONG)

        **Dùng để:** Chuyển số liệu thành danh sách hành động thực thi
        """)

    with tabs_info[7]:
        st.markdown("""
        **Phân tích động lực biến động:**
        - Đóng góp tăng/giảm theo phân khúc khách hàng
        - Đóng góp tăng/giảm theo nhóm sản phẩm
        - Tỷ lệ đóng góp ròng và theo trị tuyệt đối
        - Nhận diện nhóm kéo tăng/kéo giảm mạnh nhất

        **Dùng để:** Xác định nguyên nhân chính của biến động tổng
        """)

    with tabs_info[8]:
        st.markdown("""
        **Cảnh báo và rủi ro:**
        - Chi nhánh giảm mạnh cần kiểm soát
        - Nhóm biến động âm lớn
        - Số lượng outlier cần rà soát
        - Khách hàng ưu tiên giữ chân

        **Dùng để:** Lập danh sách hành động nhanh và ưu tiên theo rủi ro
        """)


if __name__ == "__main__":
    main()
